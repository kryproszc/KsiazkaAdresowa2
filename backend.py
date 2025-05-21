from fastapi import FastAPI, File, UploadFile, WebSocket, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Union, Optional
import uuid
import pandas as pd
import numpy as np
import asyncio
from io import BytesIO
from openpyxl import load_workbook
from calculation_method import TriangleCalculator


print("main.py")
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
CACHE = {}

class MatrixRequest(BaseModel):
    user_id: str
    paid_data: List[List[Union[str, int, float, None]]]
    paid_weights: List[List[int]]
    cl_data: List[List[Union[str, int, float, None]]]
    cl_weights: List[List[int]]
    triangle_raw: List[List[Union[str, int, float, None]]]
    cl_weights_raw: List[List[int]]
    quantiles: Optional[List[float]] = None
    nbr_samples: Optional[int] = 1000

class QuantileRequest(BaseModel):
    request_id: str
    quantiles: List[float]

@app.get("/")
async def root():
    return {"message": "Hello World"}

@app.post("/upload")
async def create_upload_files(file: UploadFile):
    print(file.filename)
    content = await file.read()
    wb = load_workbook(filename=BytesIO(content))
    ws = wb.active
    if ws:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    print(cell, end=" ")
            print("")
    return {"filename": file.filename, "size": len(content)}

@app.post("/calc/cl")
async def calc_cl(
    data: List[List[Union[str, int, float, None]]],
    selected: List[List[int]],
):
    print("selected")
    print(selected)
    df = pd.DataFrame(data).iloc[1:, 1:]
    df = df.apply(pd.to_numeric, errors="coerce")
    tri = TriangleCalculator.full_ci_calculation(df)[0]
    safe_tri = tri.replace([np.inf, -np.inf], np.nan).astype(object).where(pd.notnull(tri), None)

    matrix = [[tri.columns.name or "AY"] + list(tri.columns)]
    for idx in tri.index:
        values = list(safe_tri.loc[idx])
        while values and values[-1] is None:
            values.pop()
        row = [idx] + values
        matrix.append(row)
    return {"data": matrix}

# ---------------------
# CORE: FULL + QUANTILES
# ---------------------


@app.post("/calc/full")
async def calc_full(payload: MatrixRequest):
    print("df_triangle_raw")
    print(pd.DataFrame(payload.triangle_raw))
    df_triangle_raw = (
            pd.DataFrame(payload.triangle_raw)
            .iloc[1:, 1:]
            .apply(pd.to_numeric, errors="coerce")
        )
    df_cl_weights_raw = pd.DataFrame(payload.cl_weights_raw).astype(float)
    cl_weights = pd.DataFrame(payload.cl_weights).astype(float)

    df_np = np.array(df_triangle_raw.values, dtype=float)
    df_wagi_np = np.array(df_cl_weights_raw.values, dtype=float)
    combined_mask = (~np.isnan(df_np)) 
    #& (df_wagi_np == 1)
    mask_np = ~np.isnan(df_np)

    tri0_np = TriangleCalculator.to_incr_np(df_np)
    ctri0_np = TriangleCalculator.to_cum_np(tri0_np)
    a2a_np = TriangleCalculator.get_a2a_factors_np(ctri0_np, mask_np)
    ctri_np = TriangleCalculator.fit_triangle_from_latest_np(df_np, mask_np)
    tri_np = TriangleCalculator.to_incr_np(ctri_np)
    r_adj_np, phi_np = TriangleCalculator.full_ci_calculation_np(df_np, mask_np)

    residuals = r_adj_np[(combined_mask) & (r_adj_np != 0) & ~np.isnan(r_adj_np)].flatten()

    sim_results_np = TriangleCalculator.run_simulations_numpy(
        tri_np, residuals, mask_np, phi_np, a2a_np, nbr_samples=payload.nbr_samples
    )

    latest = np.sum(TriangleCalculator.get_latest(df_triangle_raw))
    sim_diff_np = sim_results_np - latest

    request_id = str(uuid.uuid4())
    CACHE[request_id] = {
        "sim": sim_results_np,
        "diff": sim_diff_np,
        "latest": latest
    }

    counts, bins = np.histogram(sim_results_np, bins=50)

    return {
        "message": "OK",
        "request_id": request_id,
        "triangle_shape": df_triangle_raw.shape,
        "histogram": {
            "bins": bins.tolist(),
            "counts": counts.tolist()
        }
    }

@app.post("/calc/quantiles")
async def calc_quantiles(payload: QuantileRequest):
    cached = CACHE.get(payload.request_id)
    if cached is None:
        raise HTTPException(status_code=404, detail="Nie znaleziono danych symulacji")

    sim_results_np = cached["sim"]
    sim_diff_np = cached["diff"]

    quantiles = payload.quantiles
    values_sim = np.quantile(sim_results_np, quantiles)
    values_diff = np.quantile(sim_diff_np, quantiles)

    quantile_result = {
        str(round(q, 4)): {
            "value": float(round(v1, 2)) if np.isfinite(v1) else None,
            "value_minus_latest": float(round(v2, 2)) if np.isfinite(v2) else None
        }
        for q, v1, v2 in zip(quantiles, values_sim, values_diff)
    }

    stats = {
        "mean": {
            "value": float(np.mean(sim_results_np)),
            "value_minus_latest": float(np.mean(sim_diff_np)),
        },
        "variance": {
            "value": float(np.var(sim_results_np)),
            "value_minus_latest": float(np.var(sim_diff_np)),

        },
        "std_dev": {
            "value": float(np.std(sim_results_np)),
            "value_minus_latest": float(np.std(sim_diff_np)),
        },
        "min": {
            "value": float(np.min(sim_results_np)),
            "value_minus_latest": float(np.min(sim_diff_np)),
        },
        "max": {
            "value": float(np.max(sim_results_np)),
            "value_minus_latest": float(np.max(sim_diff_np)),
        },
    }

    print("📈 stats preview:")
    for k, v in stats.items():
        print(f"  {k}: {v}")

    return {
        "quantiles": quantile_result,
        "stats": stats
    }

@app.websocket("/ws/progress")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    for i in range(1, 11):
        await websocket.send_json({"progress": i})
        if i < 10:
            await asyncio.sleep(1)
    await websocket.close()


from pydantic import BaseModel

class PercentileRequest(BaseModel):
    request_id: str
    value: float
    source: str  # 'sim' lub 'diff'

@app.post("/calc/percentile")
async def calc_percentile(payload: PercentileRequest):
    cached = CACHE.get(payload.request_id)
    if cached is None:
        raise HTTPException(status_code=404, detail="Nie znaleziono danych symulacji")

    value = payload.value
    source = payload.source

    def compute_percentile(data, val):
        sorted_data = np.sort(data)
        percentile = np.searchsorted(sorted_data, val, side='right') / len(sorted_data)
        return float(round(percentile, 6))

    if source == 'sim':
        percentile = compute_percentile(cached["sim"], value)
    elif source == 'diff':
        percentile = compute_percentile(cached["diff"], value)
    else:
        raise HTTPException(status_code=400, detail="Nieprawidłowe źródło: musi być 'sim' lub 'diff'")

    return {
        "source": source,
        "percentile": percentile
    }
